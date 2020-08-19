import openpyxl
from openpyxl.styles import Font


def make_sheet(workbook):
    finalSheet = workbook.create_sheet(title="Final", index=0)
    topValues = ["Pirkejas", "Nr.", "Dokumento Data",
                 "Suma", "Nuolaidu suma", "Pirkejo skola"]

    finalSheet.delete_rows(1, 10)
    finalSheet.append(topValues)
    return finalSheet


def format_sheet(sheet):
    for row in sheet.iter_rows(min_row=1, max_col=sheet.max_column, max_row=sheet.max_row):
        for cell in row:
            cell.font = Font(name='Calibri', size=13)

        if row[0].row != 1:
            row[2].number_format = 'yyyy.mm.dd'
            row[3].number_format = '0.00'
            row[5].number_format = '0.00'

    # formating for the final total count
    thiscell = sheet.cell(sheet.max_row, 3)
    thiscell.number_format = '0.00'

    for col in range(1, sheet.max_column+1):
        letter = openpyxl.utils.get_column_letter(col)
        width = str(sheet.cell(2, col).value)
        sheet.column_dimensions[letter].width = len(width)+10


def reformat_data(workbook, finalSheet):
    for sheet in workbook:
        if sheet.title == "Final":
            continue
        for row in sheet.iter_rows(min_row=1, max_col=sheet.max_column, max_row=sheet.max_row):
            values = []
            if row[0].value is None:
                continue
            elif row[0].value.find("iso") != -1:  # found the value from the Total row
                for cell in row:
                    if cell.value is None:
                        continue
                    else:
                        values.append(cell.value)
            elif row[0].value.find("Ex") == -1:  # didn't find the value
                continue
            else:
                for cell in row:
                    if cell.value is None:
                        continue
                    else:
                        values.append(cell.value)

            if len(values) == 5:
                values.insert(-1, None)

            finalSheet.append(values)


def process_workbook(filename):
    workbook = openpyxl.load_workbook(filename)

    # format final sheet header
    finalSheet = make_sheet(workbook)

    # extract data from all sheets and put it in final sheet
    reformat_data(workbook, finalSheet)

    # format the final sheet
    format_sheet(finalSheet)
    newName = filename.split(".")
    newfile = newName[0]+" Formated."+newName[1]
    workbook.save(newfile)

    print('Done')


file_name = 'demo 1.xlsx'
process_workbook(file_name)